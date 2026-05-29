"""
文件解析模块 —— 将上传的文件转换为 Markdown 格式文本。

支持的文件类型：
- Word (.docx)
- PDF (.pdf)
- Excel (.xlsx / .xls)
- 纯文本 (.txt)

返回干净的 Markdown 字符串，可直接拼入 LLM 提示词。
"""

from __future__ import annotations

import os
from io import BytesIO


# ---------------------------------------------------------------------------
# Word (.docx)
# ---------------------------------------------------------------------------

def parse_docx(file_bytes: bytes) -> str:
    """
    解析 Word 文档，提取段落内容并转为 Markdown。

    处理规则：
    - 标题段落（Heading 1-9）转换为对应数量的 # 前缀
    - 加粗文本用 ** 包裹
    - 列表项保留 * 前缀
    """
    from docx import Document

    doc = Document(BytesIO(file_bytes))
    lines: list[str] = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            lines.append("")
            continue

        # 处理标题
        style_name = (para.style.name or "").lower()
        if "heading" in style_name:
            # 从样式名中提取级别数字，如 "Heading 1" -> 1
            level = 1
            for token in style_name.split():
                if token.isdigit():
                    level = int(token)
                    break
            level = min(level, 6)  # Markdown 最大支持 h6
            lines.append(f"{'#' * level} {text}")
            continue

        # 处理加粗片段
        if any(run.bold for run in para.runs if run.bold):
            parts: list[str] = []
            for run in para.runs:
                run_text = run.text
                if not run_text:
                    continue
                if run.bold:
                    parts.append(f"**{run_text}**")
                else:
                    parts.append(run_text)
            text = "".join(parts)

        # 处理列表
        if para.style.name and "list" in para.style.name.lower():
            lines.append(f"* {text}")
        else:
            lines.append(text)

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# PDF (.pdf)
# ---------------------------------------------------------------------------

def parse_pdf(file_bytes: bytes) -> str:
    """
    解析 PDF 文件，逐页提取文本并以 '---' 分隔各页。

    对于扫描版 PDF（无文字层），提取结果可能为空。
    """
    from PyPDF2 import PdfReader

    reader = PdfReader(BytesIO(file_bytes))
    pages_text: list[str] = []

    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text:
            pages_text.append(text.strip())
        else:
            pages_text.append(f"（第 {i + 1} 页未提取到文字内容）")

    return "\n\n---\n\n".join(pages_text)


# ---------------------------------------------------------------------------
# Excel (.xlsx / .xls)
# ---------------------------------------------------------------------------

def parse_excel(file_bytes: bytes) -> str:
    """
    解析 Excel 工作簿，将每个工作表转换为 Markdown 表格。

    - 第一行视为表头
    - 每个工作表前附带二级标题注明表名
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            sections.append(f"## {sheet_name}\n\n（空工作表）")
            continue

        # 表头
        header = rows[0]
        col_count = len(header)
        header_cells = [str(c) if c is not None else "" for c in header]

        table_lines: list[str] = []
        table_lines.append("| " + " | ".join(header_cells) + " |")
        table_lines.append("| " + " | ".join(["---"] * col_count) + " |")

        # 数据行
        for row in rows[1:]:
            cells = []
            for i in range(col_count):
                val = row[i] if i < len(row) else None
                cells.append(str(val) if val is not None else "")
            table_lines.append("| " + " | ".join(cells) + " |")

        sections.append(f"## {sheet_name}\n\n" + "\n".join(table_lines))

    wb.close()
    return "\n\n".join(sections)


# ---------------------------------------------------------------------------
# TXT
# ---------------------------------------------------------------------------

def parse_txt(file_bytes: bytes) -> str:
    """
    解析纯文本文件。

    优先使用 utf-8 解码，失败后回退到 gbk。
    """
    for encoding in ("utf-8", "gbk", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    # 最终兜底：忽略无法解码的字节
    return file_bytes.decode("utf-8", errors="ignore")


# ---------------------------------------------------------------------------
# 统一入口
# ---------------------------------------------------------------------------

_PARSER_MAP: dict[str, callable] = {
    ".docx": parse_docx,
    ".pdf":  parse_pdf,
    ".xlsx": parse_excel,
    ".xls":  parse_excel,
    ".txt":  parse_txt,
}

SUPPORTED_EXTENSIONS = list(_PARSER_MAP.keys())


def parse_uploaded_file(file_name: str, file_bytes: bytes) -> str:
    """
    根据文件扩展名自动选择解析器，将文件内容转换为 Markdown 文本。

    参数：
        file_name: 文件名（含扩展名），用于判断文件类型
        file_bytes: 文件的原始字节内容

    返回：
        Markdown 格式的文本字符串。若解析失败则返回错误提示信息。
    """
    _, ext = os.path.splitext(file_name.lower())
    parser = _PARSER_MAP.get(ext)

    if parser is None:
        supported = ", ".join(sorted(_PARSER_MAP.keys()))
        return f"不支持的文件类型：{ext}。目前支持：{supported}"

    try:
        return parser(file_bytes)
    except Exception as e:
        return f"文件解析失败（{file_name}）：{type(e).__name__}: {e}"
